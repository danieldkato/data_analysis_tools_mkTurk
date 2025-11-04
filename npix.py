import os
import glob
import sys
import pathlib
import numbers
import re
import numpy as np
import pandas as pd
from numpy.matlib import repmat
import openpyxl
import datetime
from data_analysis_tools_mkTurk.utils_meta import *
import warnings

def generate_imro_table(length='short', parity='columnar', short_bank=0, n=384, typ=0,
    refID=0, ap_gain=500, lf_gain=250, ap_highpass=True, output_directory=None):
    """
    Generate IMRO table specifying which bank each neuropixel channel should be
    active on. 
    
    Note this function currently only produces correctly-formatted output for 
    probe-type 0; see https://billkarsh.github.io/SpikeGLX/help/imroTables/ for 
    detail. 

    Parameters
    ----------
    length : 'short'|'long', optional
        If 'short', all channels will be active on the same bank. If 'long', 
        at least two banks will include at least one active channe. The default 
        is 'short'.
    
    parity : 'columnar'|'alternating_sides', optional
        How to distribute active channels between banks if using long configuration. 
        If set to 'columnar', all channels on one side of each active bank will
        be active. If set to 'alternating_sides', active channels within a bank
        will alternate sides. Only used if `length` is set to 'long'. The default 
        is 'columnar'.
        
    short_bank : int, optional
        Which bank all channels should be active on if configuration is `length`
        is set to 'short'. The default is 0.
        
    n : int, optional
        Overall number of channels. The default is 384.

    typ : int, optional
        Probe type. The default is 0.

    refID : int, optional
        Reference ID index. 1 corresponds to tip reference. The default is 0.

    ap_gain : int, optional
        AP band gain. The default is 500.

    lf_gain : int, optional
        LF band gain. The default is 250.

    ap_highpass : bool, optional
        Whether highpass filter is applied on AP band. The default is True.

    output_directory : str, optional
        Where to save generated text file. If None, will be saved in current
        directory.

    Returns
    -------
    tuples : list
        List of n+1 tuples, where n is the number f channels on the probe. 0-th 
        tuple specifies probe-type and number of channels, while each other tuple 
        specifies the configuration for a single channel. Again see 
        https://billkarsh.github.io/SpikeGLX/help/imroTables/ for detail on 
        how each tuple is formatted. 

    """ 
    
    indices = np.arange(n) # Array of all channel indices
    ref_array = refID*np.ones(n, dtype=int)
    apgain_array = ap_gain*np.ones(n, dtype=int)
    lfgain_array = lf_gain*np.ones(n, dtype=int)
    highpass_array = ap_highpass*np.ones(n, dtype=int)    
    
    # If length is 'short', then all channels are in the same bank:
    if length == 'short':
        banks = short_bank*np.ones(n)
        parity = None
        additional_info = '_bank{}'.format(str(short_bank))
        
    # If length is 'long', then several different options for how to distribute
    # channels across different banks:
    elif length == 'long':
        
        banks = np.zeros(n)
        additional_info = '_' + parity
        
        # If all sites within a bank should be on same side of probe:
        if parity == 'columnar':
            banks[np.arange(0,n,2)+1] = 1
        
        elif parity == 'alternating_sides':
            banks[np.arange(0,n,4)+1] = 1
            banks[np.arange(0,n,4)+2] = 1
            
    banks = banks.astype(int)
    
    ch_tuples = list(zip(indices, banks, ref_array, apgain_array, lfgain_array, highpass_array))
    tuples = [(typ, n)] + ch_tuples
    
    # Define current output directory as default if necessary:
    if output_directory == None:
        output_directory = os.getcwd()

    # Create output directory if necessary:
    if not os.path.exists(output_directory):
        pathlib.Path(output_directory).mkdir(parents=True, exist_ok=True)
        
    fname = '{}{}_ref{}'.format(length, additional_info, refID)
    fpath = os.path.join(output_directory, fname)
    
    # Raise warning if overwriting existing file:
    if os.path.exists(fpath):
        response = input('Warning: text file {} already exists. Enter ''y'' to overwrite, any other key to abort.\n'.format(fpath))
        if not response.lower() == 'y':
            print('Aborting execution.\n')
            return
        
    # Write file:
    f = open(fpath, 'w')
    for t in tuples:
        f.write(str(t))
    f.write('\n')
    f.close()
    
    return tuples

    

def h5_2_ch_meta(h5path):
    zero_coords = pd.read_hdf(h5path, 'zero_coordinates')
    imro_tbl = pd.read_hdf(h5path, 'imro_table')
    imro_tbl['ch_idx_glx'] = imro_tbl.index
    
    # Assign monkey and date if possible:
    trial_df = pd.read_hdf(h5path, 'trial_params_short')
    if len(trial_df.monkey.unique())==1:
        monkey = trial_df.iloc[0].monkey
        zero_coords['monkey'] = monkey
        imro_tbl['monkey'] = monkey
        
    if len(trial_df.date.unique())==1:
        date = trial_df.iloc[0].date
        zero_coords['date'] = date
        imro_tbl['date'] = date

    return zero_coords, imro_tbl



def chs_meta_2_site_coords(zero_coords_df, imro_df, spacing=15, tip_length=175):
    """
    Compute stereotaxic coordinates of all recording sites for a given session.      

    Parameters
    ----------
    zero_coords_df : pandas.core.series.Series
        Pandas Series specifying where probe intersects pial surface of brain, in 
        stereotaxic coordinates. Should define following keys:
            
            monkey : str
                Monkey name.
                
            date : str
                Recording session date, formatted YYYY-MM-DD. 
            
            hole_id : int
                Craniotomy index. 
                
            penetration : int
                Index of penetration into craniotomy specified in `hole_id`.
                
            AP : numeric
                Anteroposterior coordinate of penetration site, in mm.     
                
            DV : numeric
                Dorsoventral coordinate of penetration site, in mm.

            ML : numeric
                Mediolateral coordinate of penetration site, in mm.
                
            Ang : numeric
                Vertical angle of probe, in degrees.
                
            HAng : numeric
                Horizontal angle of probe, in degrees.
                
            depth : numeric
                Maximum depth of probe from pial surface, in mm. 
                
            hemisphere : 'left' | 'right'
                Brain hemisphere. 
        
    imro_df : pandas.core.frame.DataFrame
        SpikeGLX IMRO table specifying channel configurations. Each row corresponds
        to a SpikeGLX channel/contact site on the proble. Should define following 
        columns:

            monkey : str
                Monkey name.
                
            date : str
                Recording session date, formatted YYYY-MM-DD. 
                
            ch_idx_glx : int
                SpikeGLX channel index. 
            
            bank : int
                Bank corresponding channel is active on in given recording session.
                
    spacing : numeric, optional
        Spacing between rows of contact sites, in micros. The default is 15.
        
    tip_length : numeric, optional
        Probe tip length, in micros. The default is 175.

    Returns
    -------
    chs_df : pandas.core.frame.DataFrame
        Dataframe of recording site coordinates. Defines the following columns:
            
            monkey : str
                Monkey name.
                
            date : str 
                Recording session date.
                
            ch_idx_glx : int
                SpikeGLX channel index. 
                
            ch_idx_depth : int
                Channel depth-order; numbering starts from most distal channel. 
                
            ap  : numpy.float
                Recording site anteroposterior coordinate, in mm. 
                
            dv  : numpy.float
                Recording site dorsoventral coordinate, in mm.
            
            mk  : numpy.float
                Recording site mediolateral coordinate, in mm.
                
            depth : numpy.float
                Channel depth from pial surface, in mm. 
    """
    
    # Initialize dataframe:
    chs_df = pd.DataFrame()
    
    # Convert zero_coords from Series to DataFrame if necessary:
    if type(zero_coords_df) == pd.Series:
        zero_coords_df = pd.DataFrame(zero_coords_df).T
    
    # Iterate over sessions:
    for zidx, zero_coords in zero_coords_df.iterrows():
        
        print(type(zero_coords))
        curr_imro_tbl = imro_df[np.array(imro_df.monkey == zero_coords.monkey) & np.array(imro_df.date== zero_coords.date)]
        curr_imro_tbl.index = curr_imro_tbl.ch_idx_glx
        curr_coords_df = get_site_coords(zero_coords, curr_imro_tbl, spacing=spacing, tip_length=tip_length)
        curr_coords_df['monkey'] = zero_coords.monkey
        curr_coords_df['date'] = zero_coords.date
        chs_df = pd.concat([chs_df, curr_coords_df], axis=0)
    
    chs_df = chs_df[['monkey', 'date', 'ch_idx_glx', 'ch_idx_depth', 'ap', 'dv', 'ml', 'depth']]
    chs_df.index = np.arange(chs_df.shape[0])
    
    return chs_df 



def get_site_coords(zero_coords, imro_tbl, spacing=20, tip_length=175):
    """
    Compute coordinates of neuropixels probe recording site. 

    Parameters
    ----------

    Returns
    -------

    """
    
    spacing = spacing/1000 # convert to mm
    
    # Get 0-coordinates:
    if np.isnan(zero_coords.HAng):
        zero_coords.HAng = 0
    depth_adjusted = (zero_coords.depth - tip_length)/1000 # Adjust for probe tip length
    
    # Compute unit vector pointing in direction of probe in ML-DV plane: 
    ehat = np.expand_dims(np.array([0,-1,0]), axis=1) # AP, ML, DV
    
    # Convert degrees to radians, angle rel. horiz.        
    angle_vert =  90 - zero_coords.Ang # Convert to relative to horizontal
    angle_vert = 2*np.pi * angle_vert/360  # Convert degrees to radians
    angle_horiz = 2*np.pi * zero_coords.HAng/360 # Convert degrees to radians
    
    # Define rotation in ML-DV plane
    Rvert = [[1,0,0],
         [0, np.cos(angle_vert), -np.sin(angle_vert)],
         [0, np.sin(angle_vert), np.cos(angle_vert)]]
    
    # Define rotation in AP-ML plane
    Rhoriz = [[np.cos(angle_horiz),-np.sin(angle_horiz),0],
     [np.sin(angle_horiz), np.cos(angle_horiz), 0],
    [0, 0, 1]]

    # Apply rotations:    
    bhat = np.matmul(Rhoriz, ehat)
    bhat = np.matmul(Rvert, bhat) # Unit vector pointing in direction of probe
    
    # Apply scale and offset:
    offset = np.expand_dims(np.array([zero_coords.AP, zero_coords.ML, zero_coords.DV]), axis=1)
    distal_coords =  offset + bhat * depth_adjusted 
    
    # Get number of channels:
    n_chans = imro_tbl.shape[0]
    Chs = np.array(imro_tbl.index)  
    
    # Get bank assignments for each channel:
    Banks = imro_tbl.bank
    
    # Compute distance of each recording site from (adjusted) tip:
    bank_length = n_chans/2*spacing 
    Distances = bank_length*Banks + spacing*((Chs - Chs % 2)/2) # < Array of recording site distances from (adjusted) tip

    # Compute 3D coordinates of each recoding site:
    B = repmat(bhat.T, n_chans, 1)
    D = np.transpose(repmat(Distances, 3, 1))
    F = repmat(distal_coords.T, n_chans, 1)
    Coords = F - np.multiply(D, B)
    
    # Save as pandas dataframe:
    coords_df = pd.DataFrame(columns=['ch_idx_glx', 'ap', 'ml', 'dv', 'depth'], index=Chs)
    coords_df['ch_idx_glx'] = Chs
    coords_df['ap'] = Coords[:,0]
    coords_df['ml'] = Coords[:,1]
    coords_df['dv'] = Coords[:,2]    
    coords_df['depth'] = depth_adjusted - D
    
    # Add channel index by depth:
    coords_df = coords_df.sort_values(by=['depth'], ascending=[False])
    coords_df['ch_idx_depth'] = np.arange(coords_df.shape[0])
    coords_df = coords_df[['ch_idx_glx', 'ch_idx_depth', 'ap', 'dv', 'ml', 'depth']]
    coords_df.index = np.arange(coords_df.shape[0])
    
    return coords_df



def extract_imro_table(metadata_path):
    """
    Extract IMRO table specifying which channels are active on which banks from
    ephys recording metadata file.

    Parameters
    ----------
    metadata_path : str
        Path to metadata file.

    Returns
    -------
    tuples : list
        List of tuples. The 0-th tuple is a pair describing the probe-type and
        overall number of channels. Each subsequent tuple describes the 
        configuration of a single channel, although the specific format depends
        on the probe-type. See https://billkarsh.github.io/SpikeGLX/help/imroTables/ 
        for detail. 

    """
    
    f = open(metadata_path,'r')
    lines = f.readlines()
    f.close()
    
    # Find lines defining IMRO table:
    imroTbl_lines = [x for x in lines if 'imroTbl' in x]
    
    # Raise error if number of lines enoding IMRO table is not exactly equal to 1:
    if len(imroTbl_lines) == 0:
        raise AssertionError('No IMRO table detected in input file {}.'.format(metadata_path))
    elif len(imroTbl_lines) > 1:
        raise AssertionError('More than one IMRO table detected in input file {}.'.format(metadata_path))
        
    # Parse line defining IMRO table into a list of strings, each encoding a tuple
    imroTbl_str = imroTbl_lines[0]
    r = re.search('\(', imroTbl_str).start() # Find index of first open parentheses
    imroTbl_str = imroTbl_str[r:] # Retain first open parens and everything after
    tuple_strs = imroTbl_str.split('(')
    tuple_strs = tuple_strs[1:] # Cut off extraneous empty string at beginning
    tuples = []
    
    # Convert strings to tuples for individual channels: 
    for tx, t in enumerate(tuple_strs):
        
        # Chop off final end parens and anything afterwards (e.g. newline):
        tail_start = re.search('\)', t).start()
        head = t[0:tail_start]
        
        if tx == 0:
            delimiter = ','
        else: 
            delimiter = ' '        
        
        # Split into strings encoding individual numbers and convert to int:
        nums = head.split(delimiter)
        nums = [int(x) for x in nums]
        
        curr_tuple = tuple(nums)
        tuples.append(curr_tuple)
        
    # Convert list of tuples to dataframe:
    imro_tbl = pd.DataFrame()
    if tuples[0][0] == 0:
        fields = ['channel', 'bank', 'ref_id', 'ap_gain', 'lf_gain', 'ap_hipass']
        
    for i, field in enumerate(fields):
        imro_tbl[field] = [x[i] for x in tuples[1:]]
        
    imro_tbl.index = imro_tbl.channel
    imro_tbl = imro_tbl.drop(columns=['channel'])
        
    return imro_tbl



def get_sess_metadata_path(base_data_path, monkey, date):
    
    data_path = get_recording_path(base_data_path, monkey, date, depth = 4)[0]
    pattern = os.path.join(data_path, '**', '*.ap.meta')
    metafiles =  glob.glob(pattern, recursive=True)
    if len(metafiles)>0:
        mfile = metafiles[0]
        session_metadata_path = os.path.join(data_path, mfile)
    else:
        warnings.warn('No .ap.meta file discovered in {}.'.format(data_path))
        session_metadata_path = None
    return session_metadata_path



def partition_adjacent_channels(coords_df, bin_size=4):
    
    # Sort sites by depth:
    coords_df = coords_df.sort_values('depth', ascending=False) # ch0 is the deepest, so sorting by depth with ascending=True would reverse channel order, which we don't want
    
    # Define bin edge indices:
    n_chan = len(coords_df.channel)
    inds = np.arange(0, n_chan+bin_size, bin_size)
    
    # Get channel ranges for each bin:
    channels = np.array(coords_df.channel)
    binned_ch_inds = [channels[inds[x]:inds[x+1]] for x in np.arange(len(inds)-1)]
    
    return binned_ch_inds


def get_session_chs(date, region=None):
    
    # Hack-y, relies on extremely hack-y date_2_chs() function; try to find a 
    # better way of doing this eventually
    
    if region == 'all':
        return np.arange(384)
    
    session_dict = date_2_chs(date)
    if session_dict is not None:
        
        channels = []
        regions = list(session_dict.keys())
        
        if region is None:
            for region in regions:
                
                if session_dict[region] is not None:
                    start_ch = session_dict[region][0]
                    stop_ch = session_dict[region][1]
                    curr_area_channels = list(np.arange(start_ch, stop_ch))
                    channels = channels + curr_area_channels
                else:   
                    continue
                
        # If requesting a particular area:               
        else:
            if session_dict[region] is not None:
                start_ch = session_dict[region][0]
                stop_ch = session_dict[region][1]            
                channels = list(np.arange(start_ch, stop_ch))
            else:
                return None
    else:
        return None
    
    # Sort and find unique:
    channels = list(np.sort(np.unique(channels)))
    
    return channels



def add_chs_by_area(df_in):
    
    # Assign some defaults:
    if 'area' not in df_in.columns:
        df_in['area'] = 'all'
    
    # Iterate over sessions x areas:
    df_out = pd.DataFrame()
    areas = df_in[['monkey', 'date', 'area']].drop_duplicates()
    for aidx, area in areas.iterrows():
        monkey = area.monkey
        date = area.date
        area = area.area
        curr_chs = session_2_chs(monkey, date, area)
        if curr_chs is None:
            curr_chs = []
        curr_chs_df = pd.DataFrame(columns=['monkey', 'date', 'area', 'ch_idx_depth'], index=np.arange(len(curr_chs)))
        curr_chs_df['monkey'] = monkey
        curr_chs_df['date'] = date
        curr_chs_df['area'] = area
        curr_chs_df['ch_idx_depth'] = curr_chs
        df_out = pd.concat([df_out, curr_chs_df], axis=0)
            
    return df_out



def session_2_chs(monkey, date=None, area=None):
    
    # Ultra-hacky; map dates to good channel ranges; find a better way of dealing
    # with this in the future:
    
    ch_lookup = {
        
        'West' : {
            
            '20230914' : {
                'IT': np.arange(280, 383),
                'WM' : np.arange(0, 279),
                'HC' : None,
                'PH' : None,
                'all' : np.arange(384)
                },
            
            '20231011' : {
                'IT': None,
                'WM' : np.arange(0, 299),
                'HC' : None,
                'PH' : np.arange(300, 384),
                'all' : np.arange(384)
                },

            '20231102' : {
                'IT': np.arange(160, 383),
                'WM' : None,
                'HC' : None,
                'PH' : np.arange(0, 159),
                'all' : np.arange(384)
                },
            
            '20231109' : {
                'IT': None,
                'WM' : np.arange(250, 383),
                'HC' : None,
                'PH' : np.arange(0, 249),
                'all' : np.arange(384)
                },        

            '20231207' : {
                'IT': np.arange(75, 384),
                'WM' : None,
                'HC' : None,
                'PH' : np.arange(0, 74),
                'all' : np.arange(384)
                },                

            '20231211' : {
                'IT': np.arange(185, 383),
                'WM' : None,
                'HC' : None,
                'PH' : np.arange(0, 184),
                'all' : np.arange(384)
                },                 

            '20240110' : {
                'IT': np.arange(285, 383),
                'WM' : np.arange(220, 285),
                'HC' : np.arange(0, 220),
                'PH' : None,
                'all' : np.arange(384)
                },            

            '20240116' : {
                'IT': np.arange(295, 383),
                'WM' : np.arange(120, 199),
                'HC' : np.arange(120, 240),
                'PH' : np.arange(0, 70),
                'all' : np.arange(384)
                },                    
            
            '20240123' : {
                'IT': np.arange(345, 383),
                'WM' : np.arange(185, 314),
                'HC' : np.arange(110, 314),
                'PH' : np.arange(0, 70),
                'all' : np.arange(384)
                },           
            
            '20240124' : {
                'IT': np.arange(250, 384),
                'WM' : np.arange(200, 250),
                'HC' : np.arange(75, 200),
                'PH' : np.arange(0, 35),
                'all' : np.arange(384)
                },           
            
            '20240130' : {
                'IT': np.arange(250, 383),
                'WM' : np.arange(40, 164),
                'HC' : np.arange(75, 210),
                'PH' : np.arange(0, 50),
                'all' : np.arange(384)
                },                    

            '20240124' : {
                'IT': np.arange(250, 383),
                'WM' : np.arange(0, 159),
                'HC' : np.arange(75, 200),
                'PH' : np.arange(0, 35),
                'all' : np.arange(384)
                },            

            '20240202' : {
                'IT': np.arange(22, 348),
                'WM' : np.arange(190, 220),
                'HC' : np.arange(80, 190),
                'PH' : None,
                'all' : np.arange(384)
                },    

            '20240207' : {
                'IT': np.arange(280, 383),
                'WM' : np.arange(230, 260),
                'HC' : np.arange(129, 235),
                'PH' : np.arange(0, 119),
                'all' : np.arange(384)
                },           

            '20240208' : {
                'IT': np.arange(265, 384),
                'WM' : np.arange(75, 199),
                'HC' : np.arange(130, 230),
                'PH' : np.arange(0, 129),
                'all' : np.arange(384)
                }, 

            '20240307' : {
                'IT': np.arange(250, 330),
                'WM' : np.arange(0, 70),
                'HC' : None,
                'PH' : np.arange(70, 250),
                'all' : np.arange(384)
                }, 

            '20240408' : {
                'IT': np.arange(285, 350),
                'WM' : np.arange(180, 285),
                'HC' : np.arange(0, 180),
                'PH' : None,
                'all' : np.arange(384)
                }, 

            '20240409' : {
                'IT': np.arange(320, 383),
                'WM' : np.arange(290, 320),
                'HC' : np.arange(0, 290),
                'PH' : None,
                'all' : np.arange(384)
                }, 

            '20240410' : {
                'IT': np.arange(285, 330),
                'WM' : np.arange(230, 285),
                'HC' : None,
                'PH' : np.arange(0, 60),
                'all' : np.arange(384)
                }, 

            '20240412' : {
                'IT': np.arange(240, 270),
                'WM' : np.arange(150, 240),
                'HC' : None,
                'PH' : np.arange(0, 150),
                'all' : np.arange(384)
                }, 

            '20240417' : {
                'IT': np.arange(275, 300),
                'WM' : np.arange(195, 275),
                'HC' : None,
                'PH' : np.arange(0, 190),
                'all' : np.arange(384)
                },         
            
            '20240418' : {
                'IT': None,
                'WM' : None,
                'HC' : None,
                'PH' : None,
                'all' : np.arange(384)
                },     
            
            '20240607' : {
                'IT': np.arange(220, 360),
                'WM' : np.arange(190, 220),
                'HC' : np.arange(65, 190),
                'PH' : np.arange(0, 65),
                'all' : np.arange(384)
                },     

            '20240718' : {
                'IT': np.arange(190, 383),
                'WM' : None,
                'HC' : None,
                'PH' : np.arange(0, 190),
                'all' : np.arange(384)
                },     

            '20240812' : {
                'IT': np.arange(260, 383),
                'WM' : np.arange(180, 260),
                'HC' : None,
                'PH' : np.arange(135, 180),
                'all' : np.arange(384)
                },     
            
            '20240816' : {
                'IT': np.arange(175, 320),
                'WM' : np.arange(14, 175),
                'HC' : None,
                'PH' : None,
                'all' : np.arange(384)
                }, 
            
            '20240920' : {
                'all' : np.arange(384),
                'TE0' : np.arange(0,180),
                'TE3' : np.arange(180, 325),
                'IT' : np.arange(0, 325)
                },

            '20240924' : {
                'all' : np.arange(384),
                'TE2' : np.arange(106, 161),
                'TE3' : np.arange(161, 328),
                'PHC' : np.arange(0, 106),
                'IT' : np.arange(106, 328),
                'MT' : np.arange(0,106)
                },

            '20240925' : {
                'all' : np.arange(384),
                'TE2' : np.arange(106, 151),
                'TE3' : np.arange(151, 328),
                'PHC' : np.arange(0, 106),
                'IT' : np.arange(106, 328),
                'MT' : np.arange(0, 106)
                },
            
            '20240927' : {
                'all' : np.arange(384),
                'TE2' : np.arange(106, 151),
                'TE3' : np.arange(151, 304),
                'PHC' : np.arange(0, 106),
                'IT' : np.arange(106, 304),
                'MT' : np.arange(0, 106)
                },
            
            '20241004' : {
                'all' : np.arange(384),
                'TE3' : np.arange(239, 353),
                'WM' : np.arange(166, 239),
                'HC' : np.arange(0, 166),
                'IT' : np.arange(239, 352),
                'MT' : np.arange(0, 165)
                },
            
            '20241010' : {
                'all' : np.arange(384),
                'TE2' : np.arange(131, 215),
                'TE3' : np.arange(215, 304),
                'PHC' : np.arange(84, 131),
                'PRH' : np.arange(0, 84),
                'IT' : np.arange(131, 304),
                'MT' : np.arange(0, 131)
                },
            
            '20241011' : {
                'all' : np.arange(384),
                'TE2' : np.arange(144, 222),
                'TE3' : np.arange(222, 324),
                'PHC' : np.arange(84, 144),
                'PRH' : np.arange(0, 84),
                'IT' : np.arange(144, 324),
                'MT' : np.arange(0, 144)
                }
            },
        
        'Bourgeois' : {

            '20250103' : {
                'IT' : np.arange(0, 263),
                'all' : np.arange(384)
                },
            
            '20250106' : {
                'IT' : np.arange(0, 254),
                'all' : np.arange(384)
                },

            '20250107' : {
                'IT' : np.arange(231, 299),
                'MT' : np.arange(0, 168),                
                'all' : np.arange(384)
                },
            
            '20250110' : {
                'IT' : np.arange(184, 310), 
                'MT' : np.arange(0, 152),
                'all' : np.arange(384) 
                },
            
            '20250310' : {
                'all' : np.arange(384),
                'unclassified' : np.arange(229, 383),
                'TE0' : np.arange(0, 229),
                'IT' : np.arange(0, 229)
                },
            
            '20250311' : {
                'all' : np.arange(384),
                'unclassified' : np.arange(180, 383),
                'TE0' : np.arange(0, 180),
                'IT' : np.arange(0, 180)
                },

            '20250328' : {
                'all' : np.arange(384),
                'unclassified' : np.arange(293, 384),
                'TE3' : np.arange(233, 292),
                'PHC' : np.arange(0, 156),
                'WM' : np.arange(156, 233),
                'IT' : np.arange(233, 292),
                'MT' : np.arange(0, 156)
                },

            '20250331' : {
                'all' : np.arange(384),
                'unclassified' : np.arange(300, 384),
                'TE3' : np.arange(236, 300),
                'PHC' : np.arange(0, 150),
                'WM' : np.arange(151, 235),
                'IT' : np.arange(236, 300),
                'MT' : np.arange(0, 150)
                },
            
            '20250417' : {
                'all' : np.arange(384) 
                },            
            
            '20250418' : {
                'all' : np.arange(384) 
                },
            
            '20250506' : {
                'all' : np.arange(384) 
                },
            
            '20250507' : {
                'all' : np.arange(384) 
                },
            
            '20250513' : {
                'all' : np.arange(384) 
                },
            
            '20250514' : {
                'all' : np.arange(384) 
                },
            
            '20250515' : {
                'all' : np.arange(384) 
                },
            
            '20250904' : {
                'all' : np.arange(384) 
                }
            
            },
        
        }
    
    if monkey not in ch_lookup.keys():
        warnings.warn('Lookup table for requested monkey {} not found.'.format(monkey))
        output = None
    else:
        monkey_dict = ch_lookup[monkey]
        
        # Return channels for specific session if requested:
        if date is None:
            output = monkey_dict
        else:
            if date not in monkey_dict.keys():
                warnings.warn('Lookup table for requested session {}, {} not found.'.format(monkey, date))
                output = None
            else:
                date_dict = monkey_dict[date] 
                
                # Return channels for specific area if requested:
                if area is None:
                    output = date_dict
                else:
                    if area not in date_dict.keys():
                        warnings.warn('Channel range for session {}, {}, area {} not found.'.format(monkey, date, area))
                        output = None
                    else:
                        output = date_dict[area]
    
    return output



def select_areas(chs_df, areas, criterion='any'):
    """
    Select channels by brain area. 

    Parameters
    ----------
    chs_df : pandas.core.frame.DataFrame
        Dataframe of channels, each row corresponding to a single channel. Must 
        define an 'areas' column, each entry of which is a list of brain areas 
        associated with the corresponding channel. Channels may be associated 
        with zero, one, or more brain areas. 
        
        
    areas : list
        List of brain areas to sample channels from.
        
    criterion : 'any' | 'all', optional
        How to select channels from multiple brain areas. If 'any', then channels
        need only be associated with one area specified in `areas` param. If 'all',
        then a channel must be associated with all brain areas in `areas` in order
        to be included in output dataframe. The default is 'any'.

    Returns
    -------
    chs_df : pandas.core.frame.DataFrame
        Filtered dataframe.

    """
    
    # Return all channels if not requesting specific areas:
    if areas == 'all' or areas == ['all']:
        return chs_df
    
    # Exclude channels not associated with any area:
    chs_df = chs_df[~chs_df.areas.isna()]
    
    # Select channels by area:
    if criterion == 'any':
        B = chs_df.apply(lambda x : len(list(set(x.areas).intersection(set(areas)))) > 0, axis=1)
    elif criterion == 'all':
        B = chs_df.apply(lambda x : len(list(set(x.areas).intersection(set(areas)))) == len(areas), axis=1)
        
    chs_df = chs_df[B]
        
    return chs_df



def exclude_multiarea_chs(chs_df, tree=None):
    """
    Excldue channels associated with more than one brain area.   
    
    Optionally, specify exceptions if brain areas are organized hierarchically 
    (e.g., allow channels to be labeled both 'IT' and 'TE0' since the latter is
     a sub-area of the former). 
    

    Parameters
    ----------
    chs_df : pandas.core.frame.DataFrame
        Dataframe of area labels, each row a channel. Should define an 'areas'
        column, each element of which is a list of area labels associated with
        the corresponding channel. Each channel may be associated with zero, one,
        or multiple brain areas.
            
    tree : dict, optional
        Dictionary specifying hierarchical organization of different brain areas. 
        Use this to allow for multiple area labels per channel if one area is 
        a sub-area of another (e.g., a channel can be in both IT and TE0 because
        TE0 is a sub-area of IT).  
        
        Each key should correspond to a "superordinate" brain area, and its value
        should be a list of "subordinate" areas. E.g.,:
    
            tree = {'IT':['TE0', 'TE2', 'TE3'], 
                    'MT':['PHC', 'HC'], 
                    'WM':['wm']}
            
        Here, area (key) 'IT' includes sub-areas (value) 'TE0', 'TE2', and 'TE3', 
        while area 'MT' includes sub-areas 'PHC' and 'HC'. The final key 'WM' 
        simply specifies an alias, evaluating to a (lowercase) synonym contained
        in a singleton list.
        
        Set to None in order to simply exclude all dataframe rows associated with
        more than one label. 
        
        
    Returns
    -------
    chs_df_hat : pandas.core.frame.DataFrame
        Filtered dataframe.

    """
    
    if tree is not None:
        root_areas = tree.keys()
    else:
        root_areas = []
    
    # Split brain areas into 'root' and 'leaf' areas: 
    R = chs_df.apply(lambda x : list(set(x.areas).intersection(root_areas)), axis=1)
    L = chs_df.apply(lambda x : list(set(x.areas).difference(root_areas)), axis=1)
    chs_df['root_areas'] = R
    chs_df['leaf_areas'] = L
    
    # Find rows associated with more than one 'root' area:
    multiroot = chs_df.apply(lambda x : len(x.root_areas)>1, axis=1)
    
    # Find rows associated with more than one 'leaf' area:
    multileaf = chs_df.apply(lambda x : len(x.leaf_areas)>1, axis=1)
    
    # Find rows where leaf area doesn't match root:
    mismatch = chs_df.apply(lambda x : len(x.root_areas)>0 and np.any([leaf not in tree[x.root_areas[0]] for leaf in x.leaf_areas]), axis=1)
    
    # Exlude rows associated with more than one root or more than one leaf:
    chs_df_hat = chs_df[~multiroot & ~multileaf & ~mismatch]
    #Ehat = E
    
    # Merge roots and leaves:
    A = chs_df_hat.apply(lambda x : x.root_areas + x.leaf_areas, axis=1)
    chs_df_hat['areas'] = A
    
    return chs_df_hat



def read_area_label_sheets(labeled_brain_areas_path = os.path.join('/', 'mnt', 'smb', 'locker', 'issa-locker', 'users', 'Dan', 'code', 'data_analysis_tools_mkTurk', 'labeled brain areas.xlsx'),
        recording_coords_path = os.path.join('/', 'mnt', 'smb', 'locker', 'issa-locker', 'users', 'Dan', 'code', 'data_analysis_tools_mkTurk', 'recording coordinate data.xlsx'),
        exclude_oob=True, exclude_multilabels=False, tree=None, flt=None):
    """
    Read channelwise brain area assignments from AG's spreadsheets.

    Parameters
    ----------
    labeled_brain_areas_path : str, optional
        Path to 'labeled brain areas' spreadsheet.
        
        
    recording_coords_path : str, optional
        Path to 'recording coordinate data' spreadsheet.
        
    exclude_oob : bool, optional
        Whether to exclude out-of-brain ('N/A') channels. The default is True.
    
    exclude_multilabels : bool, optional
        Whether to exclude channels associated with more than one brain area. 
        The default is False.

    tree : dict, optional
        Dict specifying hierarchical relationships between brain areas. Use this 
        to define exceptions to excluding channels associated with more than one 
        brain area (e.g., a channel may be in both IT and TE0 because IT contains
        TE0). If None, then setting `exclude_multiabels` to True will exclude all
        channels associated with more than one brain area. See exclude_multiarea_chs()
        docstring for detail. The default is None.

    flt : function, optional
        Boolean function used to specify subset of channels to retrieve labels for. 
        Must be applicable to a single row of a dataframe defining same columns 
        as output of read_labeled_brain_areas_sheet() and read_recording_coordinates_data_sheet(). 
        Only rows (channels) evaluating to True under flt will be included in 
        output dataframe. The default is None.

    Returns
    -------
    chs_df : pandas.core.frame.DataFrame
        Dataframe mapping channels to brain areas. Defines the following columns:
            
            ch_idx_depth : int
                Rank order of corresponding channel's depth on probe. Higher 
                indeices are more superficial. 
                
            areas : list
                List of brain areas associated with channel.

    """
    
    # Read separate Google sheets workbooks:
    wkbka_df = read_labeled_brain_areas_sheet(labeled_brain_areas_path, flt=flt) # Read workbook entitled 'labeled brain areas'
    wkbkb_df = read_recording_coordinate_data_sheet(recording_coords_path, flt=flt) # Read workbook entitled 'recording coordinate data'

    """
    # Exclude rows with no area labels:
    null_a = wkbka_df.apply(lambda x : len(x.areas)==0, axis=1)
    wkbka_df = wkbka_df[~null_a]
    
    null_b = wkbkb_df.apply(lambda x : len(x.areas)==0, axis=1)
    wkbkb_df = wkbkb_df[~null_b]
    """
    
    # Merge workbooks:
    W = [wkbka_df, wkbkb_df]
    nonempty_dfs = [w for w in W if w.shape[0]]
    
    # If the dataframes returned from 'labeled brain areas' and 'recording 
    # coordinate data' both include more than zero rows, then merge:
    if len(nonempty_dfs) == 2:
        chs_df = pd.merge(wkbka_df, wkbkb_df, on=['monkey', 'date', 'ch_idx_depth'], how='outer')

        # Replace any nan with empty list
        areas_hat_a = wkbka_df.apply(lambda x : [] if type(x.areas)==float and np.isnan(x.areas) else x.areas, axis=1)
        wkbka_df['areas'] = areas_hat_a
        
        areas_hat_b = wkbkb_df.apply(lambda x : [] if type(x.areas)==float and np.isnan(x.areas) else x.areas, axis=1)
        wkbkb_df['areas'] = areas_hat_b
        
        # Merge area labels:
        A = chs_df.apply(lambda x : x.areas_x + x.areas_y, axis=1)
        chs_df['areas'] = A
        chs_df = chs_df.drop(columns=['areas_x', 'areas_y']) 
        
    # If exactly one of the dataframes returned from 'labeled brain areas' or
    # 'recording coordinate data' includes more than zero rows, just return 
    # keep the non-empty dataframe:
    elif len(nonempty_dfs) == 1:
        chs_df = nonempty_dfs[0]
        
    # If neither dataframe returned from either 'labeled brain areas' or 'recording
    # coordinate data' includes more than zero rows, return empty dataframe:        
    else:
        warnings.warn('No channels match requested criteria; returning empty dataframe.')
        chs_df = pd.DataFrame()
        return chs_df
            
    # Optionally exclude channels associated with more than one brain area:
    if exclude_multilabels:
        chs_df = exclude_multiarea_chs(chs_df)
    
    # Optionally exclude out-of-brain channels:
    if exclude_oob:
        is_oob = chs_df.apply(lambda x : 'N/A' in x.areas, axis=1)
        chs_df = chs_df[~is_oob]
    
    # Replace empty lists with None:
    A = chs_df.apply(lambda x : x.areas if len(x.areas)>0 else ['unknown area'], axis=1)
    chs_df['areas'] = A
    
    # Return list of input paths:
    ref_paths = [labeled_brain_areas_path, recording_coords_path]
    
    return chs_df, ref_paths



def read_labeled_brain_areas_sheet(path=os.path.join('/', 'mnt', 'smb', 'locker', 'issa-locker', 'users', 'Dan', 'code', 'data_analysis_tools_mkTurk', 'labeled brain areas.xlsx'), flt=None):
    """
    Read 'labeled brain areas'  spreadsheet, which includes AG's area labels for 
    all recording sessions from Bourgeois as of 20250916, as well as several sessions
    from West.

    """
    
    # Define constants:
    n_chans = 384
    non_area_cols = ['penetration', 'date', 'configuration', 'general notes']
    
    # Load workbook:
    wb = openpyxl.load_workbook(path)
    
    # Get names of sheets with area assignments:
    sheetnames = wb.sheetnames
    suffix = ' all recordings'
    ar_shnames = [x for x in sheetnames if suffix in x]
    
    # Iterate over sheets (assume one per monkey):
    chs_df = pd.DataFrame()
    for shname in ar_shnames:
    
        # Get monkey name, initialize dataframe for current monkey:
        monkey = shname[:-len(suffix)]
        monkey_df = pd.DataFrame()
        
        # Load sheet:
        sheet = pd.read_excel(path, sheet_name=shname)
        sheet = sheet[~sheet.date.isna()] # Exclude spreadsheet rows with no date
        isnum =sheet.apply(lambda x : isinstance(x.date, numbers.Number), axis=1)
        sheet = sheet[isnum]
        
        # Convert date to str:
        sheet['date'] = sheet['date'].astype(int).astype(str) 

        # Apply any miscellaneous filters:
        if flt is not None:
            sheet = sheet[sheet.apply(flt, axis=1)]
    
        # Iterate over rows (dates) of current sheet
        for i, row in sheet.iterrows():
            
            # Initialize dataframe for current session:
            sess_df = pd.DataFrame({'monkey':monkey, 'date':str(int(row.date)), 'areas':[None]*n_chans, 'ch_idx_depth':np.arange(n_chans)}, index=np.arange(n_chans)) # assume 384 channels per session
            
            # Get brain areas recorded for current monkey:
            areas = list(set(sheet.columns).difference(set(non_area_cols)))
            
            # Iterate over brain areas (columns) of current date (row) of current monkey (sheet)
            for area in areas:
                tmp = np.array([None]*n_chans)
                ranges = re.findall('\d{1,3}-\d{1,3}', row[area]) if type(row[area])==str else []
                for rn in ranges:
                    bounds = [int(s) for s in rn.split('-')]
                    mn = min(bounds)
                    mx = min([383, max(bounds)])
                    rows = np.arange(mn, mx)
                    tmp[rows] = area
                sess_df.loc[:, area] = tmp
    
            # Merge all area labels into single 'areas'  list:
            sess_df['areas'] = sess_df.apply(lambda x : [a for a in list(x[areas]) if a is not None], axis=1)
            sess_df = sess_df.drop(columns=areas)

            # Concatenate current session with dataframe for current monkey: 
            monkey_df = pd.concat([monkey_df, sess_df], axis=0)
    
        # Concatenate current monkey with overall dataframe:
        chs_df = pd.concat([chs_df, monkey_df], axis=0)
        
    return chs_df 
    


def read_recording_coordinate_data_sheet(path=os.path.join('/', 'mnt', 'smb', 'locker', 'issa-locker', 'users', 'Dan', 'code', 'data_analysis_tools_mkTurk', 'recording coordinate data.xlsx'), flt=None):
    """
    Read 'recording coordinate data'  spreadsheet, which includes AG's area labels for 
    recording sessions from West spanning 20231017 to 20240325.

    """

    # Define constants:
    monkey = 'West' # Assume this worbook only contains area labels for West
    n_chans = 384    

    # Load sheet:     
    sheet = pd.read_excel(path, sheet_name='brain areas')
    sheet = sheet[~sheet['channel range (IT)'].isna()] # Filter by channel range (IT)

    # Format dates to yyyymmdd str:
    dates_fmt = sheet.apply(lambda x : x.date.strftime('%Y%m%d') if type(x.date)==datetime.datetime else re.search('\d{8}' ,x.date).group(), axis=1)
    sheet.loc[:, 'date'] = dates_fmt

    # Optionally apply any additional filters:
    if flt is not None:
        sheet = sheet[sheet.apply(flt, axis=1)]
    
    # Get brain area names:
    area_cols = [x for x in sheet.columns if 'channel range' in x][:-1]
    area_names = [re.search('\(\w+\)', a).group()[1:-1] for a in area_cols]
    
    # Initialize overall dataframe:
    chs_df = pd.DataFrame()
    
    # Iterate over dates (rows):
    for i, row in sheet.iterrows():
    
        # Initialize dataframe for current session:
        sess_df = pd.DataFrame({'monkey':monkey, 'date':str(int(row.date)), 'areas':[None]*n_chans, 'ch_idx_depth':np.arange(n_chans)}, index=np.arange(n_chans)) # assume 384 channels per session
        
        # Iterate over areas (columns):
        for a, area_col in enumerate(area_cols):
    
            tmp = np.array([None]*n_chans)        
    
            # Get channel ranges for current area:
            ranges = re.findall('\d{1,3}-\d{1,3}', row[area_col]) if type(row[area_col])==str else [] 
            
            # Iterate over channel ranges:
            for rn in ranges:
                bounds = [int(s) for s in rn.split('-')]
                mn = min(bounds)
                mx = min([383, max(bounds)])
                rows = np.arange(mn, mx)
                tmp[rows] = area_names[a]
            sess_df.loc[:, area_names[a]] = tmp
                
        # Merge all area labels into single 'areas'  list:
        sess_df['areas'] = sess_df.apply(lambda x : [a for a in list(x[area_names]) if a is not None], axis=1)
        sess_df = sess_df.drop(columns=area_names)
            
        chs_df = pd.concat([chs_df, sess_df], axis=0)
        
    return chs_df